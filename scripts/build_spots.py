from __future__ import annotations

import html
import json
import re
import time
from datetime import date
from datetime import time as datetime_time
from datetime import timedelta
from pathlib import Path
from urllib.parse import quote
from urllib.request import urlopen
from xml.etree import ElementTree as ET

import pdfplumber
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TOILET_KML = ROOT / "data/raw/toilets/sapporo_park_toilets.kml"
PUBLIC_FACILITY_XLSX = ROOT / "data/raw/public_facility/011006publicfacility.xlsx"
COOLING_DIR = ROOT / "data/raw/cooling"
GEOCODE_CACHE = ROOT / "data/raw/geocode_cache.json"
OUTPUT = ROOT / "data/spots.geojson"

KML_NS = {"k": "http://www.opengis.net/kml/2.2"}


def clean_text(value: str) -> str:
    value = html.unescape(value or "")
    value = re.sub(r"<br\s*/?>", "\n", value, flags=re.I)
    value = re.sub(r"<[^>]+>", "", value)
    value = re.sub(r"\n\s*\n+", "\n", value)
    return value.strip()


def normalize_key(*values: str) -> str:
    text = "|".join(values)
    text = text.replace("\u3000", " ")
    text = re.sub(r"\s+", "", text)
    text = text.replace("－", "-").replace("―", "-").replace("‐", "-")
    return text.lower()


def normalize_sapporo_address(address: str) -> str:
    address = clean_text(address).replace("\u3000", " ")
    address = re.sub(r"\s+", "", address)
    if address.startswith("北海道"):
        return address
    if address.startswith("札幌市"):
        return "北海道" + address
    return "北海道札幌市" + address


def parse_flags(description: str) -> dict[str, str | bool]:
    text = clean_text(description)
    return {
        "winter_open": "冬期開放: ◯" in text,
        "accessible": "身障者対応: ◯" in text,
        "raw_description": text,
    }


def parse_toilet_kml() -> list[dict]:
    root = ET.parse(TOILET_KML).getroot()
    features = []

    for index, placemark in enumerate(root.findall(".//k:Placemark", KML_NS), start=1):
        name = clean_text(placemark.findtext("k:name", default="", namespaces=KML_NS))
        description = placemark.findtext("k:description", default="", namespaces=KML_NS)
        coordinates = placemark.findtext(".//k:coordinates", default="", namespaces=KML_NS)
        parts = [part.strip() for part in coordinates.split(",")]
        if len(parts) < 2:
            continue

        try:
            lng = float(parts[0])
            lat = float(parts[1])
        except ValueError:
            continue

        flags = parse_flags(description)
        notes = []
        if flags["winter_open"]:
            notes.append("冬期開放")
        else:
            notes.append("冬期未開放")
        if flags["accessible"]:
            notes.append("車いす対応")

        features.append(
            {
                "type": "Feature",
                "properties": {
                    "id": f"toilet-park-{index:04d}",
                    "kind": "toilet",
                    "name": f"{name} トイレ" if name and "トイレ" not in name else name or "公園トイレ",
                    "category": "公園トイレ",
                    "address": "",
                    "hours": "",
                    "note": " / ".join(notes),
                    "winter_open": flags["winter_open"],
                    "accessible": flags["accessible"],
                    "source": "https://www.city.sapporo.jp/ryokuka/top/koueniji/benjoichiran.html",
                    "source_detail": "札幌市公園トイレマップ KML",
                    "updated": str(date.today()),
                },
                "geometry": {
                    "type": "Point",
                    "coordinates": [lng, lat],
                },
            }
        )

    return features


def format_cell(value) -> str:
    if value is None or value == "-":
        return ""
    if isinstance(value, datetime_time):
        return value.strftime("%H:%M")
    if isinstance(value, timedelta):
        total_minutes = int(value.total_seconds() // 60)
        hours, minutes = divmod(total_minutes, 60)
        return f"{hours:02d}:{minutes:02d}"
    return str(value).strip()


def load_geocode_cache() -> dict[str, list[float]]:
    if GEOCODE_CACHE.exists():
        return json.loads(GEOCODE_CACHE.read_text(encoding="utf-8"))
    return {}


def save_geocode_cache(cache: dict[str, list[float]]) -> None:
    GEOCODE_CACHE.write_text(json.dumps(cache, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def geocode_address(address: str, cache: dict[str, list[float]]) -> list[float] | None:
    address = re.sub(r"\s+.*$", "", address.replace("\u3000", " ")).strip()
    if not address:
        return None
    if address in cache:
        return cache[address]

    url = "https://msearch.gsi.go.jp/address-search/AddressSearch?q=" + quote(address)
    try:
        with urlopen(url, timeout=12) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except Exception as exc:
        print(f"geocode failed: {address} ({exc})")
        return None

    if not payload:
        print(f"geocode no match: {address}")
        cache[address] = []
        time.sleep(0.2)
        return None

    coordinates = payload[0]["geometry"]["coordinates"]
    cache[address] = [float(coordinates[0]), float(coordinates[1])]
    time.sleep(0.2)
    return cache[address]


def cool_facility_category(name: str) -> str | None:
    if "まちづくりセンター" in name:
        return None
    if "旧・" in name:
        return None
    if "図書" in name:
        return "図書館"
    if "区民センター" in name:
        return "区民センター"
    if "地区センター" in name:
        return "地区センター"
    if "市民交流プラザ" in name:
        return "市民交流施設"
    if "市民活動" in name:
        return "市民活動施設"
    if "交流センター" in name:
        return "交流センター"
    return None


def parse_public_facilities() -> list[dict]:
    workbook = load_workbook(PUBLIC_FACILITY_XLSX, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = list(next(sheet.iter_rows(min_row=2, max_row=2, values_only=True)))
    indexes = {name: headers.index(name) for name in headers if name}
    cache = load_geocode_cache()
    features = []

    for row in sheet.iter_rows(min_row=3, values_only=True):
        name = format_cell(row[indexes["施設名称"]])
        category = cool_facility_category(name)
        if not category:
            continue

        address = format_cell(row[indexes["所在地（地番）"]])
        if not address:
            continue

        coordinates = geocode_address(address, cache)
        if not coordinates:
            continue

        open_time = format_cell(row[indexes["開館時間"]])
        close_time = format_cell(row[indexes["閉館時間"]])
        hours = f"{open_time}～{close_time}" if open_time and close_time else format_cell(row[indexes["営業時間"]])
        url = format_cell(row[indexes["URL"]])

        features.append(
            {
                "type": "Feature",
                "properties": {
                    "id": f"cool-public-{len(features) + 1:04d}",
                    "kind": "cool",
                    "name": name,
                    "category": category,
                    "address": address,
                    "hours": hours,
                    "note": "札幌市公共施設一覧から抽出。開館状況は公式情報で確認してください。",
                    "source": "https://ckan.pf-sapporo.jp/dataset/public_facility",
                    "source_detail": "札幌市公共施設一覧（R7.3.31時点）",
                    "updated": str(date.today()),
                },
                "geometry": {
                    "type": "Point",
                    "coordinates": coordinates,
                },
            }
        )

    save_geocode_cache(cache)
    return features


def cooling_facility_category(name: str) -> str:
    normalized = name.replace("－", "ー").replace("―", "ー").replace("‐", "ー")
    if "図書" in normalized:
        return "図書館"
    if "区民センター" in normalized:
        return "区民センター"
    if "地区センター" in normalized:
        return "地区センター"
    if "イオン" in normalized:
        return "商業施設"
    if "マックスバリュ" in normalized or "フードセンター" in normalized:
        return "スーパー"
    if "コープさっぽろ" in normalized or "ラルズ" in normalized or "ビッグ" in normalized or "アークス" in normalized or "ラッキー" in normalized or "東光ストア" in normalized:
        return "スーパー"
    if "ツルハ" in normalized or "サツドラ" in normalized or "薬局" in normalized:
        return "ドラッグストア・薬局"
    if "博物館" in normalized or "美術館" in normalized or "資料館" in normalized or "センター" in normalized:
        return "公共・文化施設"
    return "クーリングシェルター"


def parse_cooling_shelters(existing_keys: set[str]) -> list[dict]:
    features = []
    cache = load_geocode_cache()
    pdfs = sorted(path for path in COOLING_DIR.glob("*.pdf") if path.name != "01_chuouku.pdf")

    for pdf_path in pdfs:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                for table in page.extract_tables() or []:
                    if not table or len(table) < 2:
                        continue
                    header = [clean_text(cell or "") for cell in table[0]]
                    if "施設名" not in header or "所在地" not in header:
                        continue

                    indexes = {name: i for i, name in enumerate(header)}
                    for row in table[1:]:
                        values = [clean_text(cell or "") for cell in row]
                        name = values[indexes["施設名"]]
                        address = values[indexes["所在地"]]
                        if not name or not address:
                            continue

                        address = normalize_sapporo_address(address)
                        key = normalize_key(name, address)
                        if key in existing_keys:
                            continue

                        coordinates = geocode_address(address, cache)
                        if not coordinates:
                            continue

                        hours = values[indexes.get("利用時間", -1)] if "利用時間" in indexes else ""
                        place = values[indexes.get("利用場所", -1)] if "利用場所" in indexes else ""
                        closed = values[indexes.get("休館・休業日", -1)] if "休館・休業日" in indexes else ""
                        capacity = values[indexes.get("利用可能人数\n※(最大)", -1)] if "利用可能人数\n※(最大)" in indexes else ""
                        note = values[indexes.get("備考", -1)] if "備考" in indexes else ""
                        details = []
                        if place:
                            details.append(place)
                        if capacity:
                            details.append(f"利用可能人数: {capacity}")
                        if closed:
                            details.append(f"休館・休業日: {closed}")
                        if note:
                            details.append(note)

                        features.append(
                            {
                                "type": "Feature",
                                "properties": {
                                    "id": f"cool-shelter-{len(features) + 1:04d}",
                                    "kind": "cool",
                                    "name": name,
                                    "category": cooling_facility_category(name),
                                    "address": address,
                                    "hours": hours,
                                    "note": " / ".join(details),
                                    "source": "https://www.city.sapporo.jp/kankyo/ondanka/hyperthermia/shelter.html",
                                    "source_detail": f"札幌市クーリングシェルター一覧 PDF ({pdf_path.name})",
                                    "updated": str(date.today()),
                                },
                                "geometry": {
                                    "type": "Point",
                                    "coordinates": coordinates,
                                },
                            }
                        )
                        existing_keys.add(key)

    save_geocode_cache(cache)
    return features


def main() -> None:
    public_cool_features = parse_public_facilities()
    existing_cool_keys = {
        normalize_key(feature["properties"]["name"], feature["properties"]["address"])
        for feature in public_cool_features
    }
    shelter_features = parse_cooling_shelters(existing_cool_keys)
    cool_features = shelter_features + public_cool_features
    toilet_features = parse_toilet_kml()
    features = cool_features + toilet_features
    collection = {
        "type": "FeatureCollection",
        "metadata": {
            "title": "札幌クーレットマップ データ",
            "updated": str(date.today()),
            "description": "札幌市公開情報をもとにしたアプリ用GeoJSON。涼める公共施設と公園トイレを座標付きで収録。",
            "sources": [
                {
                    "name": "札幌市公共施設一覧",
                    "url": "https://ckan.pf-sapporo.jp/dataset/public_facility",
                    "license": "CC BY 4.0",
                },
                {
                    "name": "札幌市クーリングシェルター一覧",
                    "url": "https://www.city.sapporo.jp/kankyo/ondanka/hyperthermia/shelter.html",
                    "license": "札幌市公式サイト掲載情報",
                },
                {
                    "name": "札幌市 公園トイレについて",
                    "url": "https://www.city.sapporo.jp/ryokuka/top/koueniji/benjoichiran.html",
                    "license": "札幌市オープンデータ / DATA-SMART CITY SAPPORO 掲載情報",
                }
            ],
        },
        "features": features,
    }
    OUTPUT.write_text(json.dumps(collection, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"wrote {OUTPUT} ({len(cool_features)} cool, {len(toilet_features)} toilet, {len(features)} total)")


if __name__ == "__main__":
    main()
